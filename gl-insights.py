#Format of GL
#Needs to only have 1 tab - the GL 
#Needs to have columns labeled: Date, Num, Split, Debit, Credit, Balance which are not below row 10
#dates in GL needs to be YYYY-MM-DD format, formatted "Date"
#Column B has to be where accounts are defined
#accounts need to be numbered properly
# 1xxx or 1xxxx Assets
# 2xxx or 2xxxx Liabilities
# 3xxx or 3xxxx Equity
# 4xxx or 4xxxx Revenue
# 5xxx or 5xxxx - 9xxx or 9xxxx Expense/COGS
#Excel file needs to have only 1 tab
#Dates need to be in xxxx-xx-xx format
#Needs "gl-insights-trends.xlsm" in the same folder for trend analysis 

from tkinter import * 
from tkinter import messagebox
import re
import openpyxl
from openpyxl.utils import get_column_letter
import statistics
import os
from datetime import timedelta, date
from pathlib import Path
from openpyxl.formula.translate import Translator
from tkinter import Tk   
from tkinter.filedialog import askopenfilename
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
#getting user to choose the GL file from the PC and working on it
Tk().withdraw() 
messagebox.showinfo("Locate GL", "Please locate the General Ledger file you wish to analyze (.xlsx)")
filename = askopenfilename() 
xlsx_file = Path('GL', filename)
wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True)
gl = wb_obj.active

# each entry will have entry number, date, split, and amount (+ for debit, - for credit)
class Entry:
    def __init__(self, entryNum, d, split, debit, credit, typeOfAccount):
        self.entryNum = entryNum
        self.d = d
        self.split = split
        if debit == "None":
            self.amount = credit*-1
        else:
            self.amount = debit
        self.typeOfAccount = typeOfAccount

class EntryGroup:
    def __init__(self, entries):
        self.entries = entries
        self.size = len(self.entries)
        self.entryNum = entries[0].entryNum
        self.total = 0
        self.assetAmount = 0
        self.liabilityAmount = 0
        self.expenseAmount = 0
        self.revenueAmount = 0
        self.equityAmount = 0
        for x in entries:
            self.total = self.total + x.amount
            if(x.typeOfAccount == 1):
                self.assetAmount = self.assetAmount + x.amount
            elif (x.typeOfAccount == 2):
                self.liabilityAmount = self.liabilityAmount + x.amount
            elif (x.typeOfAccount == 3):
                self.equityAmount = self.equityAmount + x.amount
            elif (x.typeOfAccount == 4):
                self.revenueAmount = self.revenueAmount + x.amount
            else :
                self.expenseAmount = self.expenseAmount + x.amount
    
class Account:
    #initialization gives account the name, and account number which is extracted from the name
    def __init__(self, name, startingBalance):
        self.name = name
        self.startingBalance = startingBalance
        self.accountNum = re.search(r'\d+', name).group()
        self.type = int(self.accountNum)    
        while(self.type >= 10):
            self.type = int(self.type / 10)
        self.entries = []
    #addEntry function adds an entry into the "entries" array
    def addEntry(self, entry):
            self.entries.append(entry) 

#this function will create the trend analysis
def trendGraph(account):
    #get all years
    years = []
    for entry in account.entries:
        y = entry.d.year
        #print(entry.d)
        if (y not in years):
            years.append(y)
    if len(years) >= 2:
        wb_obj.create_sheet(account.name)
        wb_obj[account.name]["A1"] = "Date"
        wb_obj[account.name]["B1"] = "Actual"
        wb_obj[account.name]["C1"] = "Upper Bound"
        wb_obj[account.name]["D1"] = "Lower Bound"
        row = 2
        entryCounter = 0
        curBalance = account.startingBalance
        curDate = date(years[0],1,1)
        endDate = date(years[-1],12,31)
        lastYear = date(years[-2],12,31)
        delta = timedelta(days=1)
        beforeForecastRow = 9999999
        while(curDate <= endDate):
            wb_obj[account.name]["A" + str(row)]= curDate
            wb_obj[account.name]["A" + str(row)].number_format = 'YYYY-MM-DD'
            while entryCounter < len(account.entries) and account.entries[entryCounter].d.strftime('%Y-%m-%d') == str(curDate):
                curBalance = curBalance + account.entries[entryCounter].amount
                entryCounter = entryCounter + 1
            wb_obj[account.name]["B" + str(row)]= curBalance
            wb_obj[account.name]["B" + str(row)].number_format = '$#,##0.00;-$#,##0.00'
            if curDate == lastYear:
                beforeForecastRow = row
            if row > beforeForecastRow: 
                wb_obj[account.name]["C" + str(row)].value= "=_xlfn.FORECAST.ETS($A$"+str(row)+",$B$1:$B$"+str(beforeForecastRow)+",$A$1:$A$"+str(beforeForecastRow)+",1,1)+_xlfn.FORECAST.ETS.CONFINT($A$"+str(row)+",$B$1:$B$"+str(beforeForecastRow)+",$A$1:$A$"+str(beforeForecastRow)+",0.95,1,1)"
                wb_obj[account.name]["D" + str(row)].value= "=_xlfn.FORECAST.ETS($A$"+str(row)+",$B$1:$B$"+str(beforeForecastRow)+",$A$1:$A$"+str(beforeForecastRow)+",1,1)-_xlfn.FORECAST.ETS.CONFINT($A$"+str(row)+",$B$1:$B$"+str(beforeForecastRow)+",$A$1:$A$"+str(beforeForecastRow)+",0.95,1,1)"
                wb_obj[account.name]["C" + str(row)].number_format = '$#,##0.00;-$#,##0.00'
                wb_obj[account.name]["D" + str(row)].number_format = '$#,##0.00;-$#,##0.00'      
            curDate += delta
            row+=1
        c = LineChart()
        c.title = account.name
        c.style = 12
        c.y_axis.title = "Balance"
        c.y_axis.crossAx = 500
        c.x_axis = DateAxis(crossAx=100)
        c.x_axis.number_format = 'YYYY-MM-DD'
        c.x_axis.majorTimeUnit = "days"
        c.x_axis.title = "Days"
        data = Reference(wb_obj[account.name], min_col=2, min_row=1, max_col=4, max_row=row)
        c.add_data(data, titles_from_data=True)
        dates = Reference(wb_obj[account.name], min_col=1, min_row=2, max_row=row)
        c.set_categories(dates)
        wb_obj[account.name].add_chart(c, "F1")
        wb_obj.save(filename)

    
#function to use to retrieve value from a row
def getValue(typeColumn, row):
    value = gl[typeColumn + str(row)].value
    if value == None:
        return "None"
    return value

def findColumn(word):
    for row in gl.iter_rows(min_row=1, min_col=1, max_row=10, max_col=20):
        for cell in row:
            if cell.value == word:
                return get_column_letter(cell.column)
    # raise RuntimeError
    raise RuntimeError

def processHighVariance(accountList):
    #array to return
    highVarianceEntries = []
    for x in accountList:
        #array to hold all entry amounts
        temp = []
        for e in x.entries:
            temp.append(e.amount)
        mean = statistics.mean(temp)
        stdev = statistics.stdev(temp)
        #print(str(mean)+","+str(stdev))
        for e in x.entries:
            if abs(e.amount - mean) > (3*stdev):
                if not e.entryNum in highVarianceEntries:
                    highVarianceEntries.append(e.entryNum)
    return highVarianceEntries

def processUnusualExpenses(accountList):
    unusualExpense = []
    for x in accountList:
        if len(x.entries) < 7 and x.type >= 5:
            unusualExpense.append(x.name)
    return unusualExpense

def entryToEntryGroup(entryList):
    entryGroupList = [] #array to return
    temp = []
    temp.append(entryList[0])
    i = 1
    while(i <= len(entryList)):
        if(i == len(entryList) or entryList[i].entryNum != entryList[i-1].entryNum):
            EG = EntryGroup(temp)
            entryGroupList.append(EG)
            temp.clear()
            if i != len(entryList):
                temp.append(entryList[i])
        else:
            temp.append(entryList[i])
        i = i+1
    return entryGroupList

def processEntriesDetailed(entryGroupList, invalidEntries, accountClearingEntries, transactionOffsettingEntries):
    for x in entryGroupList:
        if x.total != 0:
            invalidEntries.append(x.entryNum)
        if x.size == 2 and x.expenseAmount < 0 and x.assetAmount > 0:
            accountClearingEntries.append(x.entryNum)
        if x.size >= 3 and x.expenseAmount != 0 and x.assetAmount != 0:
            transactionOffsettingEntries.append(x.entryNum)

def main():
    try:
        #defining array to hold accounts
        accountList = []
        #array to hold all entries in order of num
        entryList =[]
        #defining columns
        entryNum = findColumn("Num")
        d = findColumn("Date")
        split = findColumn("Split")
        debit = findColumn("Debit")
        credit = findColumn("Credit")
        balance = findColumn("Balance")
        #defining max rows
        max = gl.max_row
        row = 1
        #Starting outer loop
        while row < max:
            #setting row to first row where account starts
            while gl["B" + str(row)].value == None:
                row += 1
                if row >= max:
                    break
            #print(row)
            #extracting data for each account
            account = Account(gl["B" + str(row)].value, getValue(balance,row))
            row+=1
            while 1:
                entry = Entry(getValue(entryNum,row),getValue(d,row),getValue(split,row),getValue(debit,row),getValue(credit,row), account.type)
                row += 1
                if entry.entryNum == "None":
                    break
                account.addEntry(entry)
                entryList.append(entry)
                #print("adding entry to " + str(account.name))
                #print(len(account.entries))
            accountList.append(account)
            #print("account added to accountList")
        if (len(entryList) == 0):
            raise RuntimeError
        #sorting the entryList
        entryList.sort(key=lambda entry: entry.entryNum)
        highVarianceEntries = processHighVariance(accountList) #array that holds entry numbers of high variance entries
        unusualExpenseAccounts = processUnusualExpenses(accountList) #array that holds account names of unusual expense accounts
        invalidEntries = []
        accountClearingEntries = []
        transactionOffsettingEntries = []
        entryGroupList = entryToEntryGroup(entryList)
        processEntriesDetailed(entryGroupList, invalidEntries, accountClearingEntries, transactionOffsettingEntries)
        infoMessage = ""
        for x in accountList:
            trendGraph(x)
        if(len(invalidEntries) > 0):
            infoMessage+= "Below entries do not balance (invalid):\n"
            for i in range(len(invalidEntries)):
                infoMessage+= str(invalidEntries[i])
                if (i < len(invalidEntries)-1):
                    infoMessage+= ", "
                else:
                    infoMessage+="\n\n"

        if(len(unusualExpenseAccounts) > 0):
            infoMessage+= "Below expense accounts are flagged as unusual:\n"
            for i in range(len(unusualExpenseAccounts)):
                infoMessage+= str(unusualExpenseAccounts[i])
                if (i < len(unusualExpenseAccounts)-1):
                    infoMessage+= ", "
                else:
                    infoMessage+="\n\n"

        if(len(highVarianceEntries) > 0):
            infoMessage+= "Below entries are considered high variance entries:\n"
            for i in range(len(highVarianceEntries)):
                infoMessage+= str(highVarianceEntries[i])
                if (i < len(highVarianceEntries)-1):
                    infoMessage+= ", "
                else:
                    infoMessage+="\n\n"
        
        if(len(accountClearingEntries) > 0):
            infoMessage+= "Below entries may be (expense) account clearing entries:\n"
            for i in range(len(accountClearingEntries)):
                infoMessage+= str(accountClearingEntries[i])
                if (i < len(accountClearingEntries)-1):
                    infoMessage+= ", "
                else:
                    infoMessage+="\n\n"
        
        if(len(transactionOffsettingEntries) > 0):
            infoMessage+= "Below entries may be transaction offsetting entries:\n"
            for i in range(len(transactionOffsettingEntries)):
                infoMessage+= str(transactionOffsettingEntries[i])
                if (i < len(transactionOffsettingEntries)-1):
                    infoMessage+= ", "
                else:
                    infoMessage+="\n\n"        
        if(infoMessage == ""):
            infoMessage+= "No issues found"
        messagebox.showinfo("Info", infoMessage)
    except RuntimeError:
        messagebox.showerror("ERROR","The Selected File does not follow General Ledger Formatting")

if __name__ == "__main__":
    main()